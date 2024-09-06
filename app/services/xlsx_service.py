import os.path
from functools import wraps
from os.path import exists

from flask_jwt_extended import get_jwt_identity

from app.services import expense_service, week_service, user_service
from app.models import User

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def fill_summary_data(object_dict_list,wb):
    ws = wb.create_sheet(title='Summary')
    ws.merge_cells('A1:K1')
    ws['A1'] = 'Your summary on Weekly'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = object_dict_list[0].keys()
    ws.append(list(headers))

    for row in object_dict_list:
        ws.append(list(row.values()))







def fill_week_data(week,object_dict_list,wb):
    ws = wb.create_sheet(title=f'Week {week.get_start_date()}')
    ws.merge_cells('A1:K1')
    ws['A1'] = f'Your Expenses {week.get_start_date()} - {week.get_end_date()}'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = object_dict_list[0].keys()
    ws.append(list(headers))

    for row in object_dict_list:
        ws.append(list(row.values()))

    ws.append(list(["SUM", expense_service.get_week_expenses_amount(week.id)]))


def week_to_xlsx_file(user_id_key,week_id_key):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            object_dict_list = func(*args,**kwargs)
            if not isinstance(object_dict_list, list) or not all(isinstance(item,dict) for item in object_dict_list):
                raise ValueError('Function must return a list of dictionaries')

            week = kwargs.get('week') or args[0]

            user_id = getattr(week, 'user_id')
            week_id = getattr(week, 'id')

            filename = f'user_{user_id}_week_{week_id}.xlsx'

            xlsx_path = os.path.join('app','resources','generated','xlsx')

            os.makedirs(xlsx_path, exist_ok= True)

            full_path = os.path.join(xlsx_path,filename)
            wb = Workbook()
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                std = wb['Sheet']
                wb.remove(std)
            fill_week_data(week,object_dict_list,wb)
            wb.save(full_path)
            print(f"Data exported to {full_path}")
            return object_dict_list
        return wrapper
    return decorator


def summary_to_xlsx_file():
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            object_dict_list = func(*args,**kwargs)
            if not isinstance(object_dict_list, list) or not all(isinstance(item,dict) for item in object_dict_list):
                raise ValueError('Function must return a list of dictionaries')

            user = user_service.get_user_by_email(get_jwt_identity())

            if not user:
                raise ValueError("User not found")

            filename = f'user_{user.id}_weekly_summary.xlsx'

            xlsx_path = os.path.join('app','resources','generated','xlsx')

            os.makedirs(xlsx_path, exist_ok= True)

            full_path = os.path.join(xlsx_path,filename)
            wb = Workbook()
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                std = wb['Sheet']
                wb.remove(std)
            fill_summary_data(object_dict_list,wb)


            weeks = week_service.get_all_weeks(user.id)
            for week in weeks:
                expenses = expense_service.get_all_week_expenses(week.id)
                fill_week_data(week,[expense.to_json_for_xlsx() for expense in expenses],wb)

            wb.save(full_path)
            print(f"Data exported to {full_path}")
            return object_dict_list
        return wrapper
    return decorator

